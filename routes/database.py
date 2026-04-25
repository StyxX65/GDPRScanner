"""
Database stats, disposition, export/import, admin PIN, preview, thumbnail
"""
from __future__ import annotations
import base64
from pathlib import Path
from flask import Blueprint, Response, jsonify, request
from routes import state
from app_config import _set_admin_pin, _verify_admin_pin, _admin_pin_is_set
from checkpoint import _clear_checkpoint, _DELTA_PATH
from cpr_detector import _extract_exif, _html_esc, _placeholder_svg

try:
    from gdpr_db import get_db as _get_db
    DB_OK = True
except ImportError:
    DB_OK = False
    def _get_db(*a, **kw): return None  # type: ignore[misc]

try:
    import document_scanner as _ds  # noqa: F401
    SCANNER_OK = True
except ImportError:
    SCANNER_OK = False

bp = Blueprint("database", __name__)


@bp.route("/api/db/stats")
def db_stats():
    """Return stats for the latest (or specified) scan, plus aggregate counts."""
    if not DB_OK: return jsonify({"error": "database not available"}), 503
    scan_id = request.args.get("scan_id", type=int)
    db   = _get_db()
    data = db.get_stats(scan_id) or {}
    # Add aggregate counts the Settings panel needs — query directly so they
    # are correct even if no scan has finished_at set yet
    try:
        import sqlite3 as _sq
        con = _sq.connect(db._path)
        con.row_factory = _sq.Row
        data["total_items"]   = con.execute("SELECT COUNT(*) FROM flagged_items").fetchone()[0]
        data["flagged_items"] = data["total_items"]
        data["total_scans"]   = con.execute("SELECT COUNT(*) FROM scans").fetchone()[0]
        data["finished_scans"]= con.execute("SELECT COUNT(*) FROM scans WHERE finished_at IS NOT NULL").fetchone()[0]
        if not data.get("flagged_count"):
            data["flagged_count"] = data["total_items"]
        if not data.get("total_scanned"):
            data["total_scanned"] = con.execute("SELECT COALESCE(SUM(total_scanned),0) FROM scans").fetchone()[0]
        con.close()
    except Exception:
        data.setdefault("total_items",  0)
        data.setdefault("flagged_items", 0)
        data.setdefault("total_scans",   0)
    return jsonify(data)


@bp.route("/api/db/trend")
def db_trend():
    """Return scan history for trend chart (last 20 scans)."""
    if not DB_OK: return jsonify({"error": "database not available"}), 503
    n = request.args.get("n", default=20, type=int)
    return jsonify(_get_db().get_trend(n))


@bp.route("/api/db/scans")
def db_scans():
    """List recent completed scans."""
    if not DB_OK: return jsonify({"error": "database not available"}), 503
    return jsonify(_get_db().scans_list())


@bp.route("/api/db/sessions")
def db_sessions():
    """List scan sessions (grouped concurrent scans), newest first."""
    if not DB_OK: return jsonify([])
    return jsonify(_get_db().get_sessions())


@bp.route("/api/db/subject", methods=["POST"])
def db_subject_lookup():
    """Find all items containing a given CPR number.
    Body: {cpr: "DDMMYY-XXXX"}
    The CPR is hashed before querying -- never stored in plaintext.
    """
    if not DB_OK: return jsonify({"error": "database not available"}), 503
    data = request.get_json() or {}
    cpr  = data.get("cpr", "").strip().replace("-", "").replace(" ", "")
    if not cpr:
        return jsonify({"error": "cpr required"}), 400
    items = _get_db().lookup_data_subject(cpr)
    return jsonify({"count": len(items), "items": items})


@bp.route("/api/db/overdue")
def db_overdue():
    """Return items older than the retention threshold.

    Query params:
        years            int, default 5
        fiscal_year_end  MM-DD string, e.g. 12-31 (omit for rolling window)
        scan_id          int (omit for latest scan)
    """
    if not DB_OK: return jsonify({"error": "database not available"}), 503
    years           = request.args.get("years", default=5, type=int)
    fiscal_year_end = request.args.get("fiscal_year_end", default=None)
    scan_id         = request.args.get("scan_id", type=int)
    try:
        from gdpr_db import overdue_cutoff
        cutoff = overdue_cutoff(years, fiscal_year_end)
        items  = _get_db().get_overdue_items(years, scan_id, fiscal_year_end)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({
        "count":          len(items),
        "cutoff_date":    cutoff,
        "cutoff_mode":    "fiscal" if fiscal_year_end else "rolling",
        "fiscal_year_end": fiscal_year_end,
        "years":          years,
        "items":          items,
    })


@bp.route("/api/db/disposition", methods=["POST"])
def db_set_disposition():
    """Set a compliance disposition on a flagged item.
    Body: {item_id, status, legal_basis?, notes?, reviewed_by?}
    Status values: unreviewed | retain-legal | retain-legitimate | retain-contract |
                   delete-scheduled | deleted | personal-use
    """
    if not DB_OK: return jsonify({"error": "database not available"}), 503
    data = request.get_json() or {}
    item_id = data.get("item_id", "")
    if not item_id:
        return jsonify({"error": "item_id required"}), 400
    _get_db().set_disposition(
        item_id,
        status      = data.get("status", "unreviewed"),
        legal_basis = data.get("legal_basis", ""),
        notes       = data.get("notes", ""),
        reviewed_by = data.get("reviewed_by", ""),
    )
    return jsonify({"status": "saved"})


@bp.route("/api/db/disposition/bulk", methods=["POST"])
def db_set_disposition_bulk():
    """Set the same disposition on multiple items at once.
    Body: {item_ids: [...], status, legal_basis?, notes?, reviewed_by?}
    """
    if not DB_OK: return jsonify({"error": "database not available"}), 503
    data     = request.get_json() or {}
    item_ids = data.get("item_ids", [])
    status   = data.get("status", "")
    if not item_ids or not status:
        return jsonify({"error": "item_ids and status required"}), 400
    db = _get_db()
    for iid in item_ids:
        db.set_disposition(iid, status,
                           legal_basis=data.get("legal_basis", ""),
                           notes=data.get("notes", ""),
                           reviewed_by=data.get("reviewed_by", ""))
    return jsonify({"saved": len(item_ids)})


@bp.route("/api/db/disposition/<item_id>")
def db_get_disposition(item_id):
    """Get the current disposition for an item."""
    if not DB_OK: return jsonify({"error": "database not available"}), 503
    d = _get_db().get_disposition(item_id)
    return jsonify(d or {"status": "unreviewed"})


@bp.route("/api/db/flagged")
def db_flagged_items():
    """Return flagged items from the most recent completed scan session.
    Used by the read-only viewer to load results without an active SSE connection.
    Respects viewer_scope.role stored in the session for scoped tokens.
    """
    if not DB_OK: return jsonify([])
    from flask import session as _session
    scope     = _session.get("viewer_scope", {})
    role_filt = scope.get("role", "") if isinstance(scope, dict) else ""
    # user may be a list of emails (current) or a legacy single string
    raw_user  = scope.get("user", "") if isinstance(scope, dict) else ""
    if isinstance(raw_user, list):
        user_filt = set(e.lower() for e in raw_user if e)
    else:
        user_filt = {raw_user.lower()} if raw_user else set()
    ref_scan_id = request.args.get("ref", type=int)
    items = _get_db().get_session_items(ref_scan_id=ref_scan_id)
    # Normalise JSON-encoded columns the same way scan_engine does for SSE cards
    import json as _json
    out = []
    for row in items:
        if role_filt and row.get("user_role", "") != role_filt:
            continue
        if user_filt and (row.get("account_id", "") or "").lower() not in user_filt:
            continue
        row["special_category"] = _json.loads(row.get("special_category") or "[]") if isinstance(row.get("special_category"), str) else row.get("special_category", [])
        row["exif"] = _json.loads(row.get("exif_json") or "{}") if isinstance(row.get("exif_json"), str) else row.get("exif", {})
        row.pop("exif_json", None)
        out.append(row)
    return jsonify(out)


@bp.route("/api/db/related/<item_id>")
def db_related_items(item_id):
    """Return flagged items from the same session sharing at least one CPR hash."""
    if not DB_OK:
        return jsonify([])
    ref = request.args.get("ref", type=int)
    import json as _json
    out = []
    for row in _get_db().get_related_items(item_id, ref_scan_id=ref):
        row["special_category"] = _json.loads(row.get("special_category") or "[]") if isinstance(row.get("special_category"), str) else row.get("special_category", [])
        row["exif"] = _json.loads(row.get("exif_json") or "{}") if isinstance(row.get("exif_json"), str) else row.get("exif", {})
        row.pop("exif_json", None)
        out.append(row)
    return jsonify(out)


@bp.route("/api/db/deletion_log")
def db_deletion_log():
    """Return the deletion audit log.
    Query params: limit (int, default 500), reason (str filter)
    """
    if not DB_OK: return jsonify({"error": "database not available"}), 503
    limit  = request.args.get("limit", default=500, type=int)
    reason = request.args.get("reason", default=None)
    rows   = _get_db().get_deletion_log(limit=limit, reason=reason)
    stats  = _get_db().deletion_log_stats()
    return jsonify({"stats": stats, "entries": rows})


@bp.route("/api/db/reset", methods=["POST"])
def db_reset():
    """Reset the database and clear in-memory scan results.
    Requires {confirm: "yes", pin: "<admin_pin>"} in request body.
    """
    data = request.get_json() or {}
    if data.get("confirm") != "yes":
        return jsonify({"error": "confirm=yes required"}), 400
    if _admin_pin_is_set():
        pin = data.get("pin", "")
        if not _verify_admin_pin(pin):
            return jsonify({"error": "incorrect_pin"}), 403
    if not DB_OK:
        return jsonify({"error": "database not available"}), 503
    try:
        _get_db().reset()
        state.flagged_items = []
        state.scan_meta = {}
        _clear_checkpoint()
        if _DELTA_PATH.exists():
            _DELTA_PATH.unlink()
        return jsonify({"ok": True, "message": "Database reset. All scan results cleared."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@bp.route("/api/admin/pin", methods=["GET"])
def admin_pin_status():
    """Return whether an admin PIN has been set."""
    return jsonify({"pin_set": _admin_pin_is_set()})


@bp.route("/api/admin/pin", methods=["POST"])
def admin_pin_set():
    """Set or change the admin PIN.
    Body: {current_pin: "..", new_pin: ".."}
    If no PIN is currently set, current_pin is not required.
    """
    data = request.get_json() or {}
    new_pin = data.get("new_pin", "").strip()
    if not new_pin:
        return jsonify({"error": "new_pin required"}), 400
    if _admin_pin_is_set():
        if not _verify_admin_pin(data.get("current_pin", "")):
            return jsonify({"error": "incorrect_pin"}), 403
    _set_admin_pin(new_pin)
    return jsonify({"ok": True})


@bp.route("/api/db/export")
def db_export():
    """Export the database to a structured ZIP and return it as a download.
    The ZIP contains 8 JSON files (see ScanDB.export_db for details).
    CPR numbers are stored as SHA-256 hashes only — never in plaintext.
    Thumbnails are stripped to keep the download small.  (#11)
    """
    if not DB_OK:
        return jsonify({"error": "database not available"}), 503
    import tempfile, datetime as _dt
    try:
        ts  = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"gdpr_export_{ts}.zip"
        with tempfile.NamedTemporaryFile(suffix=".zip", delete=False) as tf:
            tmp = Path(tf.name)
        try:
            _get_db().export_db(tmp)
            data = tmp.read_bytes()
        finally:
            try: tmp.unlink()
            except Exception: pass
        return Response(
            data,
            mimetype="application/zip",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "detail": traceback.format_exc()}), 500


@bp.route("/api/db/import", methods=["POST"])
def db_import():
    """Import a previously exported ZIP archive into the database.  (#11)

    Multipart form:
        file    — the export ZIP
        mode    — "merge" (default) or "replace"
        confirm — must be "yes" when mode == "replace"
    """
    if not DB_OK:
        return jsonify({"error": "database not available"}), 503
    import tempfile
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "no file uploaded"}), 400
    mode    = request.form.get("mode", "merge")
    confirm = request.form.get("confirm", "")
    if mode == "replace" and confirm != "yes":
        return jsonify({"error": "confirm=yes required for replace mode"}), 400
    try:
        tmp = Path(tempfile.mktemp(suffix=".zip", prefix="gdpr_import_"))
        f.save(str(tmp))
        result = _get_db().import_db(tmp, mode=mode)
        tmp.unlink(missing_ok=True)
        return jsonify({"ok": True, "mode": mode, "imported": result})
    except (ValueError, FileNotFoundError) as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@bp.route("/api/preview/<item_id>")
def get_preview(item_id):
    """Return a preview URL or HTML for a flagged item."""
    source_type = request.args.get("source_type", "")
    account_id  = request.args.get("account_id", "me") or "me"

    # Local and SMB file sources — re-read file and render preview
    if source_type in ("local", "smb"):
        item_meta = next((x for x in state.flagged_items if x.get("id") == item_id), {})
        full_path = item_meta.get("full_path", "")
        name      = item_meta.get("name", "")
        ext       = Path(name).suffix.lower() if name else ""

        if not full_path:
            return jsonify({"error": "File path not available — rescan to enable preview"})

        if source_type == "smb":
            return jsonify({
                "type":  "info",
                "html":  f"<p style='color:var(--muted);font-size:12px'>SMB preview requires re-reading the file over the network. Open the file directly: <code>{full_path}</code></p>",
            })

        try:
            file_path = Path(full_path).expanduser()
            if not file_path.exists():
                return jsonify({"error": f"File not found: {full_path}"})

            size = file_path.stat().st_size

            # Images — return as data URI
            if ext in {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}:
                import base64 as _b64
                mime = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
                        "gif": "image/gif", "webp": "image/webp", "bmp": "image/bmp"}.get(ext.lstrip("."), "image/jpeg")
                data = _b64.b64encode(file_path.read_bytes()).decode()
                _exif = item_meta.get("exif") or _extract_exif(file_path.read_bytes(), name)
                exif_html = ""
                if _exif:
                    rows = []
                    if _exif.get("gps"):
                        g = _exif["gps"]
                        rows.append(f'<tr><td>📍 GPS</td><td><a href="{g["maps_url"]}" target="_blank" style="color:#7ec8d0">{g["lat"]}, {g["lon"]}</a></td></tr>')
                    if _exif.get("author"):
                        rows.append(f'<tr><td>👤 Author</td><td>{_html_esc(_exif["author"])}</td></tr>')
                    if _exif.get("datetime"):
                        rows.append(f'<tr><td>📅 Date</td><td>{_html_esc(_exif["datetime"])}</td></tr>')
                    if _exif.get("device"):
                        rows.append(f'<tr><td>📷 Device</td><td>{_html_esc(_exif["device"])}</td></tr>')
                    for field, val in (_exif.get("pii_fields") or {}).items():
                        if field not in ("Artist",):
                            rows.append(f'<tr><td>{_html_esc(field)}</td><td>{_html_esc(str(val)[:200])}</td></tr>')
                    if rows:
                        exif_html = ('<details style="margin:8px 12px;font-size:11px">'
                                     '<summary style="cursor:pointer;color:#888">EXIF data</summary>'
                                     '<table style="border-collapse:collapse;width:100%;margin-top:6px">'
                                     + "".join(f'<tr style="border-top:1px solid #333"><td style="padding:4px 8px;color:#888;width:120px;white-space:nowrap">{r.split("</td><td>")[0].replace("<tr><td>","")}</td><td style="padding:4px 8px;word-break:break-all">{r.split("</td><td>")[1].replace("</td></tr>","")}</td></tr>' for r in rows)
                                     + '</table></details>')
                html = f'<div style="text-align:center;padding:12px"><img src="data:{mime};base64,{data}" style="max-width:100%;max-height:60vh;border-radius:6px"></div>{exif_html}'
                return jsonify({"type": "html", "html": html})

            # Text-based files — render with highlighted CPR numbers
            if ext in {".txt", ".csv", ".eml", ".md", ".log", ".xml", ".json", ".html", ".htm"}:
                if size > 2 * 1024 * 1024:
                    return jsonify({"error": "File too large for inline preview (>2 MB)"})
                raw = file_path.read_bytes().decode("utf-8", errors="replace")
                import html as _html, re as _re
                escaped = _html.escape(raw[:50000])
                escaped = _re.sub(
                    r"(\d{6}[-\s]?\d{4})",
                    r'<mark style="background:#ff444455;color:#ff8888;border-radius:2px">\1</mark>',
                    escaped
                )
                html_out = (
                    '<pre style="font-family:var(--mono);font-size:11px;white-space:pre-wrap;'
                    'word-break:break-all;padding:12px;color:var(--text);line-height:1.6">'
                    + escaped + "</pre>"
                )
                return jsonify({"type": "html", "html": html_out})

            # PDF — render first 5 pages as text using pdfplumber
            if ext == ".pdf":
                if size > 20 * 1024 * 1024:
                    return jsonify({"error": "File too large for preview (>20 MB)"})
                if SCANNER_OK:
                    try:
                        import pdfplumber as _plumber, io as _io, html as _h
                        pages_html = []
                        with _plumber.open(_io.BytesIO(file_path.read_bytes())) as pdf:
                            total = len(pdf.pages)
                            for i, page in enumerate(pdf.pages[:5]):
                                text = page.extract_text() or ""
                                if not text.strip():
                                    text = f"[Page {i+1}: image-only or OCR required]"
                                import re as _re
                                escaped = _re.sub(
                                    r"(\d{6}[-\s]?\d{4})",
                                    r'<mark style="background:#ff444455;color:#ff8888;border-radius:2px">\1</mark>',
                                    _h.escape(text)
                                )
                                pages_html.append(
                                    f'<div style="border-bottom:1px solid #333;padding:10px 0;margin-bottom:8px">'
                                    f'<div style="font-size:9px;color:#666;margin-bottom:4px">Page {i+1}</div>'
                                    f'<pre style="font-size:11px;white-space:pre-wrap;word-break:break-all;margin:0;line-height:1.6">{escaped}</pre>'
                                    f'</div>'
                                )
                        note = f'<div style="font-size:10px;color:#666;padding:6px 0">Showing {min(5,total)} of {total} page(s)</div>' if total > 5 else ""
                        html_out = f'<div style="padding:10px">{note}{"".join(pages_html)}</div>'
                        return jsonify({"type": "html", "html": html_out})
                    except Exception:
                        pass
                html_out = (
                    f'<div style="padding:24px;text-align:center;font-family:sans-serif">'
                    f'<div style="font-size:40px">📄</div>'
                    f'<div style="font-size:13px;font-weight:600;margin:8px 0">{_html_esc(name)}</div>'
                    f'<div style="font-size:11px;color:var(--muted)">{round(size/1024,1)} KB</div>'
                    f'<div style="margin-top:12px;font-size:11px;color:var(--muted)">{_html_esc(full_path)}</div>'
                    f'</div>'
                )
                return jsonify({"type": "html", "html": html_out})

            # Word/Excel/CSV — render content or show metadata
            if SCANNER_OK and ext in {".xlsx", ".xlsm", ".csv"}:
                try:
                    import html as _hh, re as _re, io as _io
                    if ext == ".csv":
                        raw = file_path.read_bytes().decode("utf-8", errors="replace")
                        rows = [r for r in raw.splitlines()[:50]]
                        table_rows = ""
                        for i, row in enumerate(rows):
                            cols = row.split(",")
                            style = "background:#2a2a2a" if i % 2 == 0 else ""
                            cells = "".join(f'<td style="padding:3px 8px;border:1px solid #333;max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{_hh.escape(str(c)[:80])}</td>' for c in cols)
                            table_rows += f'<tr style="{style}">{cells}</tr>'
                        html_out = f'<div style="padding:8px;overflow-x:auto"><table style="border-collapse:collapse;font-size:11px;color:var(--text)">{table_rows}</table></div>'
                    else:
                        import openpyxl as _xl
                        wb = _xl.load_workbook(_io.BytesIO(file_path.read_bytes()), read_only=True, data_only=True)
                        tabs = []
                        for sheet_name in wb.sheetnames[:3]:
                            ws = wb[sheet_name]
                            table_rows = ""
                            for i, row in enumerate(ws.iter_rows(max_row=50, values_only=True)):
                                style = "background:#2a2a2a" if i % 2 == 0 else ""
                                cells = "".join(
                                    f'<td style="padding:3px 8px;border:1px solid #333;max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">'
                                    f'{_hh.escape(str(c)[:80]) if c is not None else ""}</td>'
                                    for c in row
                                )
                                table_rows += f'<tr style="{style}">{cells}</tr>'
                            tabs.append(
                                f'<div style="margin-bottom:12px">'
                                f'<div style="font-size:10px;color:#888;margin-bottom:4px">📋 {_hh.escape(sheet_name)}</div>'
                                f'<div style="overflow-x:auto"><table style="border-collapse:collapse;font-size:11px;color:var(--text)">{table_rows}</table></div>'
                                f'</div>'
                            )
                        html_out = '<div style="padding:8px">' + "".join(tabs) + '</div>'
                    return jsonify({"type": "html", "html": html_out})
                except Exception:
                    pass

            if SCANNER_OK and ext in {".docx", ".doc"}:
                try:
                    import io as _io, html as _hh, re as _re
                    from docx import Document as _Doc
                    doc = _Doc(_io.BytesIO(file_path.read_bytes()))
                    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()][:80]
                    text = "\n".join(paragraphs)
                    escaped = _re.sub(
                        r"(\d{6}[-\s]?\d{4})",
                        r'<mark style="background:#ff444455;color:#ff8888;border-radius:2px">\1</mark>',
                        _hh.escape(text)
                    )
                    html_out = f'<div style="padding:12px"><pre style="font-size:11px;white-space:pre-wrap;word-break:break-all;line-height:1.7">{escaped}</pre></div>'
                    return jsonify({"type": "html", "html": html_out})
                except Exception:
                    pass

            html_out = (
                f'<div style="padding:24px;text-align:center;font-family:sans-serif">'
                f'<div style="font-size:40px">📄</div>'
                f'<div style="font-size:13px;font-weight:600;margin:8px 0">{_html_esc(name)}</div>'
                f'<div style="font-size:11px;color:var(--muted)">{round(size/1024,1)} KB · {ext.upper().lstrip(".")} file</div>'
                f'<div style="margin-top:12px;font-size:11px;color:var(--muted)">{_html_esc(full_path)}</div>'
                f'</div>'
            )
            return jsonify({"type": "html", "html": html_out})

        except PermissionError:
            return jsonify({"error": f"Permission denied: {full_path}"})
        except Exception as e:
            return jsonify({"error": str(e)})

    if not state.connector:
        return jsonify({"error": "not authenticated"}), 401

    item_meta = next((x for x in state.flagged_items if x.get("id") == item_id), {})
    drive_id  = item_meta.get("drive_id", "")

    try:
        if source_type == "email":
            uid = account_id
            try:
                msg = state.connector._get(
                    f"/{'me' if uid == 'me' else 'users/' + uid}/messages/{item_id}",
                    {"$select": "subject,from,receivedDateTime,body"}
                )
            except Exception as e:
                return jsonify({"error": f"Could not load email: {e}"})

            sender   = msg.get("from", {}).get("emailAddress", {})
            from_str = f"{sender.get('name', '')} &lt;{sender.get('address', '')}&gt;"
            date_str = (msg.get("receivedDateTime") or "")[:10]
            body_html = msg.get("body", {}).get("content", "") or ""
            content_type = msg.get("body", {}).get("contentType", "text")
            import html as _html
            if content_type == "text":
                body_html = "<pre style='white-space:pre-wrap;font-family:sans-serif'>" + _html.escape(body_html) + "</pre>"

            att_list = item_meta.get("attachments", [])
            att_html = ""
            if att_list:
                def _att_row(a):
                    cpr_badge = f'<span class="att-cpr">{a["cpr_count"]} CPR</span>' if a["cpr_count"] else ''
                    name_esc  = _html.escape(a["name"])
                    return f'<div class="att-row"><span class="att-name">{name_esc}</span>{cpr_badge}</div>'
                rows = "".join(_att_row(a) for a in att_list)
                att_html = f"""
<div class="att-section">
  <div class="att-header">📎 Attachments ({len(att_list)})</div>
  {rows}
</div>"""

            page = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
  *, *::before, *::after {{ box-sizing: border-box; max-width: 100%; }}
  html, body {{ margin: 0; padding: 0; overflow-x: hidden; }}
  body {{ font-family: -apple-system, sans-serif; font-size: 13px; padding: 12px 16px;
         background: #fff; color: #111; word-break: break-word; }}
  img {{ max-width: 100% !important; height: auto !important; }}
  table {{ max-width: 100% !important; table-layout: fixed; word-break: break-word; }}
  .hdr {{ border-bottom: 1px solid #eee; margin-bottom: 12px; padding-bottom: 10px; }}
  .hdr-row {{ color: #555; font-size: 12px; margin-bottom: 3px; }}
  .hdr-row b {{ color: #111; }}
  .att-section {{ margin-top: 16px; border-top: 1px solid #eee; padding-top: 10px; }}
  .att-header {{ font-size: 12px; font-weight: 600; color: #555; margin-bottom: 6px; }}
  .att-row {{ display: flex; align-items: center; gap: 8px; font-size: 12px;
              padding: 4px 0; border-bottom: 1px solid #f0f0f0; }}
  .att-name {{ flex: 1; color: #333; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
  .att-cpr {{ background: #fff0f0; color: #c00; font-size: 11px; padding: 1px 6px;
              border-radius: 10px; font-weight: 600; white-space: nowrap; }}
  ::-webkit-scrollbar {{ width: 4px; height: 4px; }}
  ::-webkit-scrollbar-track {{ background: transparent; }}
  ::-webkit-scrollbar-thumb {{ background: #aaa; border-radius: 2px; }}
  * {{ scrollbar-width: thin; scrollbar-color: #aaa transparent; }}
</style></head><body>
<div class="hdr">
  <div class="hdr-row"><b>From:</b> {from_str}</div>
  <div class="hdr-row"><b>Date:</b> {date_str}</div>
  <div class="hdr-row"><b>Subject:</b> {_html.escape(msg.get('subject', '(no subject)'))}</div>
</div>
{body_html}{att_html}
</body></html>"""
            return jsonify({"type": "html", "html": page})

        else:
            # OneDrive / SharePoint / Teams — use Graph's embed preview API
            preview_url = None
            errors = []

            endpoints_to_try = []
            if drive_id:
                endpoints_to_try.append(f"/drives/{drive_id}/items/{item_id}/preview")
            uid = account_id
            if uid and uid != "me":
                endpoints_to_try.append(f"/users/{uid}/drive/items/{item_id}/preview")
            endpoints_to_try.append(f"/me/drive/items/{item_id}/preview")

            for ep in endpoints_to_try:
                try:
                    data = state.connector._post(ep, {})
                    preview_url = data.get("getUrl") or data.get("postUrl")
                    if preview_url:
                        break
                except Exception as e:
                    errors.append(str(e))

            if preview_url:
                return jsonify({"type": "iframe", "url": preview_url})
            return jsonify({"error": "No preview available for this file type. " + "; ".join(errors[:1])})

    except Exception as e:
        return jsonify({"error": str(e)})


@bp.route("/api/thumb")
def thumb():
    """Fallback thumbnail for non-image files."""
    name = request.args.get("name", "file")
    ext  = Path(name).suffix.lower()
    svg_b64 = _placeholder_svg(ext, name)
    data = base64.b64decode(svg_b64)
    return Response(data, mimetype="image/svg+xml",
                    headers={"Cache-Control": "public, max-age=3600"})
